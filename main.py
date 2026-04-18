from fastapi import FastAPI, Request
import openpyxl
import os
import re

app = FastAPI()

# -----------------------------
# SAFE MESSAGE EXTRACTION
# -----------------------------
def get_message_text(data):
    try:
        return data["entry"][0]["changes"][0]["value"]["messages"][0]["text"]["body"]
    except:
        return ""


# -----------------------------
# UNIVERSAL ATTENDANCE PARSER
# -----------------------------
def parse_employee_line(line):
    """
    Accepts ANY of these formats:
    NR009 GURMEET = P+3
    NR009 GURMEET P+3
    NR009 GURMEET - P+3
    NR009 GURMEET: P+3
    NR009 GURMEET P 3
    NR009 GURMEET A
    NR009 GURMEET P--
    """

    # Extract employee code (NRxxx)
    match = re.match(r"(NR\d+)", line)
    if not match:
        return None

    code = match.group(1)

    # Extract attendance part (P+3, A, P--, etc.)
    # Look for patterns like P+3, P--, A, P, etc.
    att_match = re.search(r"(P\+\d+|P-\-+|P-|P|A-\-+|A)", line.replace(" ", ""))
    if not att_match:
        return None

    value = att_match.group(1)

    # Normalize value
    v = value.strip()

    # Case 1: P+3
    if "+" in v:
        parts = v.split("+")
        att = parts[0] or "P"
        ot = parts[1] if len(parts) > 1 else "0"

    # Case 2: P-- or P- or A--
    elif "-" in v:
        parts = v.split("-")
        att = parts[0] or "A"
        ot = "0"

    # Case 3: Only P or A
    else:
        att = v
        ot = "0"

    return code, att, ot


# -----------------------------
# SAFE ATTENDANCE EXTRACTION
# -----------------------------
def extract_attendance(text):
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    # Find DATE line safely
    date_candidates = [l for l in lines if "DATE" in l.upper() or "/" in l]
    if not date_candidates:
        return {"error": "No date found", "raw_text": text}

    date_line = date_candidates[0]

    # Company name = line after date (if exists)
    try:
        company_line = lines[lines.index(date_line) + 1]
    except:
        company_line = "UNKNOWN COMPANY"

    # Extract employees using universal parser
    employees = []
    for l in lines:
        parsed = parse_employee_line(l)
        if parsed:
            employees.append(parsed)

    return {
        "date": date_line,
        "company": company_line,
        "employees": employees
    }


# -----------------------------
# WRITE TO EXCEL TEMPLATE
# -----------------------------
def write_to_excel(extracted):
    print("Starting Excel write...")

    try:
        wb = openpyxl.load_workbook("template.xlsx")
    except:
        print("ERROR: template.xlsx not found in Railway container")
        return

    ws = wb["Attendance"]

    # Extract date number (e.g., 18 from DATE: 18/04/2026)
    date_line = extracted["date"]
    date_num = int(date_line.split(":")[1].strip().split("/")[0])
    print(f"Looking for date column: {date_num}")

    # AUTO-DETECT DATE COLUMN
    date_col = None
    date_row = None

    for row in range(1, 50):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue

            clean = str(cell_value).replace("\n", "").replace("\t", "").strip()

            if clean == str(date_num):
                date_row = row
                date_col = col
                print(f"FOUND DATE {date_num} at row {row}, col {col}")
                break
        if date_col:
            break

    if not date_col:
        print("ERROR: Date column not found in template (scanned rows 1–50)")
        return

    print(f"Using row {date_row} and column {date_col} for date {date_num}")

    pa_col = date_col
    ot_col = date_col + 1

    # WRITE ATTENDANCE
    for code, att, ot in extracted["employees"]:

        # Find employee row
        emp_row = None
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and str(cell_val).strip() == code:
                emp_row = row
                break

        if not emp_row:
            print(f"Employee {code} not found in template")
            continue

        print(f"Writing for {code}: ATT={att}, OT={ot}")

        ws.cell(row=emp_row, column=pa_col).value = att
        ws.cell(row=emp_row, column=ot_col).value = ot

    wb.save("attendance_output.xlsx")
    print("Excel updated successfully! File saved as attendance_output.xlsx")


# -----------------------------
# WEBHOOK ENDPOINT
# -----------------------------
@app.post("/webhook")
async def receive_whatsapp(request: Request):
    data = await request.json()

    text = get_message_text(data)

    print("RAW MESSAGE RECEIVED:")
    print(text)

    extracted = extract_attendance(text)

    print("EXTRACTED DATA:")
    print(extracted)

    write_to_excel(extracted)

    return {"status": "received", "extracted": extracted}


# -----------------------------
# WEBHOOK VERIFICATION
# -----------------------------
@app.get("/webhook")
async def verify(request: Request):
    params = request.query_params
    mode = params.get("hub.mode")
    challenge = params.get("hub.challenge")
    token = params.get("hub.verify_token")

    VERIFY_TOKEN = os.getenv("VERIFY_TOKEN")

    if mode == "subscribe" and token == VERIFY_TOKEN:
        return int(challenge)

    return {"error": "Invalid verification token"}
